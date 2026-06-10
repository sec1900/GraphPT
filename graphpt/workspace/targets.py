"""目标文本解析与分类，支持 Excel/CSV 批量导入。"""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from graphpt.workspace import _workspace_target_dirs, _workspace_targets_dir

_URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)
_IP_RE = re.compile(r"\b\d{1,3}(\.\d{1,3}){3}(/\d{1,2})?\b")
_DOMAIN_RE = re.compile(r"\b([a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}\b")

_VALID_TLDS = frozenset()  # 不再限制 TLD，允许所有域名通过

_SCAN_URL_RE = re.compile(r"https?://[^\s,;\"'<>]+", re.IGNORECASE)
_SCAN_IP_RE = re.compile(r"\b(\d{1,3}(?:\.\d{1,3}){3})(?:/(\d{1,2}))?\b")
_SCAN_DOMAIN_RE = re.compile(r"\b((?:[a-zA-Z0-9](?:[a-zA-Z0-9-]*[a-zA-Z0-9])?\.)+([a-zA-Z]{2,}))\b")
_TARGET_SCHEMA_FILENAMES = ("targets.yaml", "targets.yml", "targets.json")
_SCHEMA_TARGET_KEYS = ("domains", "subdomains", "urls", "ips", "cidrs", "companies")
_COMPOUND_PUBLIC_SUFFIXES = frozenset({
    "co.uk", "org.uk", "gov.uk", "ac.uk",
    "com.cn", "net.cn", "org.cn", "gov.cn",
    "com.hk", "com.tw", "com.au", "net.au",
    "co.jp", "or.jp", "com.br",
})


def extract_targets_from_text(text: str) -> list[str]:
    """从自然语言文本中扫描提取域名/IP/URL，用 TLD 白名单过滤误判。"""
    seen: set[str] = set()
    results: list[str] = []

    for m in _SCAN_URL_RE.finditer(text):
        val = m.group(0).rstrip(".,;)>")
        low = val.lower()
        if low not in seen:
            seen.add(low)
            results.append(val)

    for m in _SCAN_IP_RE.finditer(text):
        val = m.group(0)
        low = val.lower()
        if low not in seen:
            seen.add(low)
            results.append(val)

    for m in _SCAN_DOMAIN_RE.finditer(text):
        val = m.group(1)
        low = val.lower()
        if low not in seen:
            seen.add(low)
            results.append(val)

    return results


def parse_structured_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """智能解析 xlsx/csv/txt，识别表头列名映射，分类提取目标/黑名单/描述。"""
    _TARGET_KEYS = {"目标", "target", "targets", "url", "urls", "域名", "domain", "domains", "ip", "ips", "地址", "address", "host"}
    _BLACKLIST_KEYS = {"黑名单", "排除", "blacklist", "exclude", "excluded", "block"}
    _DESC_KEYS = {"描述", "说明", "任务", "要求", "description", "desc", "task", "note", "notes", "备注", "instruction"}

    result: dict[str, Any] = {"targets": [], "blacklist": [], "description": "", "raw": []}

    if filename.lower().endswith(".txt"):
        text = file_bytes.decode("utf-8", errors="replace")
        result["raw"] = _split_target_entries(text)
        return result
    if filename.lower().endswith(".csv"):
        text = file_bytes.decode("utf-8", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    elif filename.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl_not_installed")
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c).strip() if c is not None else "" for c in row])
        wb.close()
    else:
        raise ValueError("unsupported_file_type")

    if not rows:
        return result

    # 尝试识别表头行
    header = rows[0]
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header):
        key = str(cell).strip().lower()
        if key in _TARGET_KEYS:
            col_map[idx] = "targets"
        elif key in _BLACKLIST_KEYS:
            col_map[idx] = "blacklist"
        elif key in _DESC_KEYS:
            col_map[idx] = "description"

    if col_map:
        desc_parts: list[str] = []
        for row in rows[1:]:
            for idx, category in col_map.items():
                if idx >= len(row):
                    continue
                val = str(row[idx]).strip()
                if not val:
                    continue
                if category == "targets":
                    result["targets"].append(val)
                elif category == "blacklist":
                    result["blacklist"].append(val)
                elif category == "description":
                    desc_parts.append(val)
        result["description"] = "\n".join(desc_parts)
    else:
        # 无法识别表头 → 回退到提取全部非空单元格
        for row in rows:
            for cell in row:
                val = str(cell).strip()
                if val:
                    result["raw"].append(val)

    return result


def _split_target_entries(text: str) -> list[str]:
    entries: list[str] = []
    for line in text.splitlines():
        for part in line.split(","):
            s = part.strip()
            if s:
                entries.append(s)
    return entries


def _dedupe_strings(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        item = str(value or "").strip()
        if not item:
            continue
        key = item.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def _empty_targets_schema() -> dict[str, Any]:
    return {
        "schema_version": 1,
        "targets": {
            "domains": [],
            "subdomains": [],
            "urls": [],
            "ips": [],
            "cidrs": [],
            "companies": [],
        },
        "scope": {
            "mode": "none",
            "whitelist": [],
            "blacklist": [],
        },
        "notes": [],
    }


def _split_domain_kind(value: str) -> str:
    host = str(value or "").strip().strip(".").lower()
    if not host:
        return "domains"
    labels = [part for part in host.split(".") if part]
    if len(labels) <= 2:
        return "domains"
    tail2 = ".".join(labels[-2:])
    tail3 = ".".join(labels[-3:]) if len(labels) >= 3 else ""
    if tail2 in _COMPOUND_PUBLIC_SUFFIXES:
        return "domains" if len(labels) <= 3 else "subdomains"
    if tail3 in _COMPOUND_PUBLIC_SUFFIXES:
        return "domains" if len(labels) <= 4 else "subdomains"
    return "subdomains"


def _base_domain(value: str) -> str:
    host = str(value or "").strip().strip(".").lower()
    if not host:
        return ""
    labels = [part for part in host.split(".") if part]
    if len(labels) <= 2:
        return host
    tail2 = ".".join(labels[-2:])
    if tail2 in _COMPOUND_PUBLIC_SUFFIXES:
        return ".".join(labels[-3:]) if len(labels) >= 3 else host
    tail3 = ".".join(labels[-3:]) if len(labels) >= 3 else ""
    if tail3 in _COMPOUND_PUBLIC_SUFFIXES:
        return ".".join(labels[-4:]) if len(labels) >= 4 else host
    return ".".join(labels[-2:])


def _add_unique_value(targets: dict[str, list[str]], bucket: str, value: str) -> None:
    item = str(value or "").strip()
    if not item:
        return
    if item not in targets[bucket]:
        targets[bucket].append(item)


def _classify_host_value(result: dict[str, list[str]], host: str) -> None:
    normalized = str(host or "").strip().strip(".").lower()
    if not normalized:
        return
    if _IP_RE.fullmatch(normalized):
        _add_unique_value(result, "ips", normalized)
        return
    if not _DOMAIN_RE.fullmatch(normalized):
        return
    bucket = _split_domain_kind(normalized)
    if bucket == "subdomains":
        _add_unique_value(result, "subdomains", normalized)
        root_domain = _base_domain(normalized)
        if root_domain:
            _add_unique_value(result, "domains", root_domain)
        return
    _add_unique_value(result, "domains", normalized)


def _classify_target_entries(entries: list[str]) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {
        "domains": [],
        "subdomains": [],
        "urls": [],
        "ips": [],
        "cidrs": [],
        "companies": [],
    }
    seen: set[str] = set()

    for entry in entries:
        item = str(entry or "").strip()
        if not item:
            continue
        lower = item.casefold()
        if lower in seen:
            continue

        if _URL_RE.match(item):
            _add_unique_value(result, "urls", item)
            host = urlparse(item).hostname or ""
            _classify_host_value(result, host)
            seen.add(lower)
            continue

        if _IP_RE.fullmatch(item):
            if "/" in item:
                result["cidrs"].append(item)
            else:
                result["ips"].append(item)
            seen.add(lower)
            continue

        if _DOMAIN_RE.fullmatch(item):
            _classify_host_value(result, item)
            seen.add(lower)
            continue

        result["companies"].append(item)
        seen.add(lower)

    return result


def classify_targets(text: str) -> dict[str, list[str]]:
    """将自由文本分类为 ips / domains / subdomains / urls / companies。"""
    classified = _classify_target_entries(_split_target_entries(text))
    return {
        "ips": classified["ips"] + classified["cidrs"],
        "domains": classified["domains"],
        "subdomains": classified["subdomains"],
        "urls": classified["urls"],
        "companies": classified["companies"],
    }


def build_targets_schema(
    data: str | list[str] | dict[str, Any],
    *,
    scope_whitelist: list[str] | None = None,
    scope_blacklist: list[str] | None = None,
    scope_mode: str = "blacklist",
    notes: list[str] | None = None,
) -> dict[str, Any]:
    schema = _empty_targets_schema()
    if isinstance(data, dict) and "targets" in data:
        return normalize_targets_schema(data)
    if isinstance(data, dict):
        targets = dict(data)
        cidrs: list[str] = []
        ips: list[str] = []
        for item in list(targets.get("ips", []) or []):
            value = str(item or "").strip()
            if not value:
                continue
            if "/" in value:
                cidrs.append(value)
            else:
                ips.append(value)
        schema["targets"]["domains"] = _dedupe_strings(list(targets.get("domains", []) or []))
        schema["targets"]["subdomains"] = _dedupe_strings(list(targets.get("subdomains", []) or []))
        schema["targets"]["urls"] = _dedupe_strings(list(targets.get("urls", []) or []))
        schema["targets"]["ips"] = _dedupe_strings(ips)
        schema["targets"]["cidrs"] = _dedupe_strings(list(targets.get("cidrs", []) or []) + cidrs)
        schema["targets"]["companies"] = _dedupe_strings(list(targets.get("companies", []) or []))
    else:
        entries = _split_target_entries(data) if isinstance(data, str) else [str(v).strip() for v in (data or []) if str(v).strip()]
        classified = _classify_target_entries(entries)
        for key in _SCHEMA_TARGET_KEYS:
            schema["targets"][key] = _dedupe_strings(classified.get(key, []))

    schema["scope"]["mode"] = str(scope_mode or "none").strip().lower() or "none"
    schema["scope"]["whitelist"] = _dedupe_strings(list(scope_whitelist or []))
    schema["scope"]["blacklist"] = _dedupe_strings(list(scope_blacklist or []))
    schema["notes"] = _dedupe_strings(list(notes or []))
    return schema


def normalize_targets_schema(raw: Any) -> dict[str, Any]:
    schema = _empty_targets_schema()
    if not isinstance(raw, dict):
        return schema

    targets_raw = raw.get("targets", raw)
    if isinstance(targets_raw, dict):
        normalized = build_targets_schema(targets_raw)
        schema["targets"] = normalized["targets"]

    scope_raw = raw.get("scope", {})
    if isinstance(scope_raw, dict):
        mode = str(scope_raw.get("mode", "none")).strip().lower()
        schema["scope"]["mode"] = mode if mode in {"none", "whitelist", "blacklist"} else "none"
        schema["scope"]["whitelist"] = _dedupe_strings(list(scope_raw.get("whitelist", []) or []))
        schema["scope"]["blacklist"] = _dedupe_strings(list(scope_raw.get("blacklist", []) or []))

    schema["notes"] = _dedupe_strings(list(raw.get("notes", []) or []))
    raw_version = raw.get("schema_version")
    try:
        schema["schema_version"] = max(1, int(raw_version or 1))
    except (TypeError, ValueError):
        schema["schema_version"] = 1
    return schema


def summarize_targets_schema(schema: dict[str, Any]) -> dict[str, Any]:
    normalized = normalize_targets_schema(schema)
    targets = normalized["targets"]
    counts = {key: len(list(targets.get(key, []) or [])) for key in _SCHEMA_TARGET_KEYS}
    items = (
        list(targets.get("urls", []) or [])
        + list(targets.get("domains", []) or [])
        + list(targets.get("subdomains", []) or [])
        + list(targets.get("ips", []) or [])
        + list(targets.get("cidrs", []) or [])
        + list(targets.get("companies", []) or [])
    )
    return {
        "total": sum(counts.values()),
        "classified": counts,
        "items": items,
        "preview": items,
        "scope": normalized["scope"],
        "notes": normalized["notes"],
    }


def _parse_targets_schema_text(text: str) -> dict[str, Any]:
    stripped = text.strip()
    if not stripped:
        return _empty_targets_schema()
    try:
        return normalize_targets_schema(json.loads(stripped))
    except (TypeError, ValueError):
        pass
    try:
        import yaml as _yaml

        loaded = _yaml.safe_load(stripped)
        return normalize_targets_schema(loaded)
    except (ImportError, AttributeError, ValueError):
        return _empty_targets_schema()


def _targets_schema_candidates(workspace_root: Path) -> list[Path]:
    out: list[Path] = []
    for targets_dir in _workspace_target_dirs(workspace_root):
        for name in _TARGET_SCHEMA_FILENAMES:
            candidate = targets_dir / name
            if candidate not in out:
                out.append(candidate)
    return out


def get_targets_schema_info(workspace_root: Path) -> dict[str, Any]:
    for candidate in _targets_schema_candidates(workspace_root):
        if not candidate.is_file():
            continue
        schema = _parse_targets_schema_text(candidate.read_text(encoding="utf-8", errors="replace"))
        return {
            "source": "schema",
            "schema": schema,
            "schema_file": str(candidate),
        }
    return {
        "source": "empty",
        "schema": _empty_targets_schema(),
        "schema_file": "",
    }


def load_targets_schema(workspace_root: Path) -> dict[str, Any]:
    return normalize_targets_schema(get_targets_schema_info(workspace_root)["schema"])


def write_target_files(workspace_root: Path, classified: dict[str, list[str]]) -> list[str]:
    """将分类结果写入统一 schema 文件，返回已写文件列表。"""
    targets_dir = _workspace_targets_dir(workspace_root)
    targets_dir.mkdir(parents=True, exist_ok=True)
    schema = build_targets_schema(classified)
    target_file = targets_dir / "targets.yaml"
    target_file.write_text(json.dumps(schema, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return [str(target_file)]


def parse_excel_targets(file_bytes: bytes, filename: str) -> list[str]:
    """解析 .xlsx / .csv / .txt 文件，提取所有非空条目。"""
    entries: list[str] = []

    if filename.lower().endswith(".txt"):
        text = file_bytes.decode("utf-8", errors="replace")
        entries.extend(_split_target_entries(text))
    elif filename.lower().endswith(".csv"):
        text = file_bytes.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            for cell in row:
                s = cell.strip()
                if s:
                    entries.append(s)
    elif filename.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl_not_installed")
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        s = str(cell).strip()
                        if s:
                            entries.append(s)
        wb.close()
    else:
        raise ValueError("unsupported_file_type")

    return entries


def merge_target_files(workspace_root: Path, new_entries: list[str]) -> dict[str, Any]:
    """读取已有 schema，合并新条目后写回统一 schema。返回摘要。"""
    existing = load_targets_schema(workspace_root)
    incoming = build_targets_schema(new_entries)
    merged = _empty_targets_schema()
    for key in _SCHEMA_TARGET_KEYS:
        merged["targets"][key] = _dedupe_strings(
            list(existing["targets"].get(key, []) or []) + list(incoming["targets"].get(key, []) or [])
        )
    merged["scope"] = dict(existing.get("scope", {}))
    merged["notes"] = _dedupe_strings(list(existing.get("notes", []) or []) + list(incoming.get("notes", []) or []))
    written = write_target_files(workspace_root, merged)
    summary = summarize_targets_schema(merged)
    summary["files_written"] = written
    summary["schema_file"] = written[0] if written else ""
    return summary
